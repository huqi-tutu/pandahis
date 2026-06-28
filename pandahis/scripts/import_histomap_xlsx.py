#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将《历史图谱数据.xlsx》导入 histomap MySQL（或生成 SQL 文件）。

用法:
  python scripts/import_histomap_xlsx.py --xlsx "C:/Users/user/Downloads/历史图谱数据.xlsx" --sql-out import_batch.sql
  python scripts/import_histomap_xlsx.py --xlsx ... --mysql-host 127.0.0.1 --mysql-user root --mysql-password x --mysql-db histomap

产品约定（与后端 schema 一致）:
  - 一级地域保持 18 条 canonical；Excel「一级文明归属」文本映射到 civilization_l1.id。
  - 史略主键 historical_box.id 使用 Excel「史略 ID」列；business_code 与之相同便于查询。
  - 矩阵 category_key 由「史略分类」映射到 junji/shichen/dianzhi/shilue/minlu。
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import defaultdict
from pathlib import Path


def sql_escape(s: str | None) -> str:
    if s is None:
        return "NULL"
    return "'" + str(s).replace("\\", "\\\\").replace("'", "''") + "'"


def parse_year(val) -> int | None:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return int(val)
    t = str(val).strip().replace(" ", "").replace("约", "").replace("前", "-")
    if t == "" or t == "-":
        return None
    if re.fullmatch(r"-?\d+", t):
        return int(t)
    m = re.match(r"^(-?\d+)", t)
    if m:
        return int(m.group(1))
    return None


SHILUE_KIND_TO_CATEGORY = {
    "君纪": "junji",
    "士臣": "shichen",
    "典制": "dianzhi",
    "事略": "shilue",
    "民录": "minlu",
    "论著": "lunzhu",
}

# Excel「一级文明归属」常见写法 -> civilization_l1.id（与 data.sql 18 条一致）
CIV_TEXT_TO_ID = {
    "华夏": 1,
    "华夏文明": 1,
    "朝鲜": 2,
    "日本": 3,
    "东南亚": 4,
    "中亚": 5,
    "北亚": 6,
    "北亚游牧": 6,
    "印度": 7,
    "南亚": 7,
    "波斯两河": 8,
    "西亚": 8,
    "阿拉伯": 8,
    "南欧": 9,
    "地中海古典": 9,
    "东欧": 10,
    "西欧": 11,
    "北欧": 12,
    "埃及": 13,
    "北非": 13,
    "西非": 14,
    "东非": 15,
    "中美洲": 16,
    "中美": 16,
    "北美": 17,
    "安第斯": 18,
    "南美": 18,
}


def civ_id_for(text: str | None) -> int:
    if not text:
        return 1
    t = str(text).strip()
    return CIV_TEXT_TO_ID.get(t, 1)


def sheet_rows(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    return rows[0], rows[1:]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, type=Path)
    ap.add_argument("--sql-out", type=Path, default=None)
    ap.add_argument("--mysql-host", default=None)
    ap.add_argument("--mysql-port", type=int, default=3306)
    ap.add_argument("--mysql-user", default=None)
    ap.add_argument("--mysql-password", default="")
    ap.add_argument("--mysql-db", default=None)
    args = ap.parse_args()

    try:
        import openpyxl
    except ImportError:
        print("需要: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    # 勿用 read_only=True：部分工作簿维度元数据异常时 iter_rows 只返回表头一行，导致「全量」实际为 0 条。
    wb = openpyxl.load_workbook(args.xlsx, read_only=False, data_only=True)
    stmts: list[str] = []
    excel_box_ids: list[str] = []

    def emit(sql: str):
        stmts.append(sql.rstrip(";") + ";")

    # --- 文明 ---
    if "文明" in wb.sheetnames:
        h, data = sheet_rows(wb["文明"])
        for row in data:
            if not row or not row[0]:
                continue
            name, code = row[0], row[1] if len(row) > 1 else None
            # 仅更新 code/tab_image（id 由种子数据固定）
            emit(
                f"UPDATE civilization_l1 SET code={sql_escape(str(code).strip() if code else None)} "
                f"WHERE display_name={sql_escape(str(name).strip())} LIMIT 1;"
            )

    # --- 王朝 ---
    dyn_rows = []
    if "王朝" in wb.sheetnames:
        h, data = sheet_rows(wb["王朝"])
        for row in data:
            if not row or not row[0]:
                continue
            dname, civn, sy, ey, did = (list(row) + [None] * 5)[:5]
            cid = civ_id_for(civn)
            syi, eyi = parse_year(sy), parse_year(ey)
            did_s = str(did).strip() if did else f"dyn_{abs(hash((str(dname), cid))) % 10000000000}"
            dyn_rows.append((did_s, cid, str(dname).strip(), syi, eyi, str(sy) if sy else None, str(ey) if ey else None))
            emit(
                "INSERT INTO historical_dynasty (id, civilization_l1_id, name, start_year, end_year, start_year_raw, end_year_raw, sort_order, status) "
                f"VALUES ({sql_escape(did_s)}, {cid}, {sql_escape(str(dname).strip())}, "
                f"{syi if syi is not None else 'NULL'}, {eyi if eyi is not None else 'NULL'}, "
                f"{sql_escape(str(sy)) if sy else 'NULL'}, {sql_escape(str(ey)) if ey else 'NULL'}, 0, 1) "
                "ON DUPLICATE KEY UPDATE name=VALUES(name), civilization_l1_id=VALUES(civilization_l1_id), "
                "start_year=VALUES(start_year), end_year=VALUES(end_year), start_year_raw=VALUES(start_year_raw), end_year_raw=VALUES(end_year_raw);"
            )

    dynasty_by_name_civ: dict[tuple[str, int], str] = {(str(a[2]).strip(), int(a[1])): a[0] for a in dyn_rows}

    # --- 帝王 -> historical_unit ---
    unit_by_name: dict[str, str] = {}
    if "帝王" in wb.sheetnames:
        h, data = sheet_rows(wb["帝王"])
        for row in data:
            if not row or not row[0]:
                continue
            uid = str(row[0]).strip()
            dynasty_name = str(row[1] or "").strip()
            ruler = str(row[2] or "").strip()
            uname = str(row[3] or "").strip()
            temple = row[4] if len(row) > 4 else None
            era_t = row[5] if len(row) > 5 else None
            sy, ey = parse_year(row[6]) if len(row) > 6 else None, parse_year(row[7]) if len(row) > 7 else None
            dur = int(row[8]) if len(row) > 8 and row[8] is not None else 0
            imp = int(row[9]) if len(row) > 9 and row[9] is not None else None
            notes = str(row[10]).strip() if len(row) > 10 and row[10] else None
            civ_t = str(row[11]).strip() if len(row) > 11 and row[11] else ""
            cid = civ_id_for(civ_t)
            dyn_id = dynasty_by_name_civ.get((dynasty_name, cid))
            dyn_sql = sql_escape(dyn_id) if dyn_id else "NULL"
            syv = sy if sy is not None else 0
            eyv = ey if ey is not None else syv
            if dur == 0 and eyv != syv:
                dur = max(1, eyv - syv)
            if dur == 0:
                dur = 1
            unit_by_name[uname] = uid
            unit_by_name[ruler] = uid
            emit(
                "INSERT INTO historical_unit (id, name, ruler_name, dynasty_name, dynasty_id, era_name, temple_name, era_title, "
                "civilization_l1_id, start_year, end_year, duration_years, importance_level, core_topics_json, summary, notes, status) "
                f"VALUES ({sql_escape(uid)}, {sql_escape(uname)}, {sql_escape(ruler) if ruler else 'NULL'}, "
                f"{sql_escape(dynasty_name) if dynasty_name else 'NULL'}, {dyn_sql}, "
                f"{sql_escape(str(era_t)) if era_t else 'NULL'}, "
                f"{sql_escape(str(temple)) if temple else 'NULL'}, NULL, {cid}, {syv}, {eyv}, {dur}, "
                f"{imp if imp is not None else 'NULL'}, '[]', {sql_escape(notes or '')}, NULL, 1) "
                "ON DUPLICATE KEY UPDATE name=VALUES(name), ruler_name=VALUES(ruler_name), dynasty_name=VALUES(dynasty_name), "
                "dynasty_id=VALUES(dynasty_id), era_name=VALUES(era_name), temple_name=VALUES(temple_name), "
                "civilization_l1_id=VALUES(civilization_l1_id), start_year=VALUES(start_year), end_year=VALUES(end_year), "
                "duration_years=VALUES(duration_years), importance_level=VALUES(importance_level), summary=VALUES(summary), notes=VALUES(notes);"
            )

    # --- 史略 ---
    if "史略" in wb.sheetnames:
        h, data = sheet_rows(wb["史略"])
        for row in data:
            if not row or not row[0]:
                continue
            bid = str(row[0]).strip()
            title = str(row[1] or "").strip()
            blurb = str(row[2] or "").strip()[:64] if len(row) > 2 and row[2] else None
            prio = str(row[3] or "").strip() if len(row) > 3 and row[3] else None
            prio_r = str(row[4] or "").strip() if len(row) > 4 and row[4] else None
            skind = str(row[5] or "").strip() if len(row) > 5 and row[5] else ""
            subject_name = str(row[6] or "").strip() if len(row) > 6 and row[6] else ""
            unit_id = unit_by_name.get(subject_name) or unit_by_name.get(title) or ""
            if not unit_id:
                unit_id = "huaxia_song_shenzong"
            cat = SHILUE_KIND_TO_CATEGORY.get(skind, "shilue")
            sy, ey = parse_year(row[9]) if len(row) > 9 else None, parse_year(row[10]) if len(row) > 10 else None
            src_line = str(row[11] or "").strip() if len(row) > 11 else ""
            excerpt = str(row[12] or "").strip() if len(row) > 12 else ""
            detail = str(row[13] or "").strip() if len(row) > 13 else ""
            flash = str(row[14] or "").strip() if len(row) > 14 else ""
            pro = str(row[15] or "").strip() if len(row) > 15 else ""
            ref = {"title": "史料原文", "items": []}
            if src_line or excerpt:
                ref["items"].append({"work": src_line[:128], "chapter": "", "excerpt": excerpt, "url": ""})
            ref_json = json.dumps(ref, ensure_ascii=False)
            imp = 3
            if prio and prio.upper().startswith("P0"):
                imp = 5
            elif prio and prio.upper().startswith("P1"):
                imp = 4
            syv = sy if sy is not None else 0
            eyv = ey if ey is not None else syv
            emit(
                "INSERT INTO historical_box (id, business_code, unit_id, subject_unit_id, dynasty_id, title, category_key, shilue_kind, blurb, "
                "start_year, end_year, importance_level, priority_code, priority_reason, status, detail_md, detail_md_flash, detail_md_pro, original_ref_json) "
                f"VALUES ({sql_escape(bid)}, {sql_escape(bid)}, {sql_escape(unit_id)}, NULL, NULL, "
                f"{sql_escape(title)}, {sql_escape(cat)}, {sql_escape(skind) if skind else 'NULL'}, "
                f"{sql_escape(blurb) if blurb else 'NULL'}, {syv}, {eyv}, {imp}, "
                f"{sql_escape(prio) if prio else 'NULL'}, {sql_escape(prio_r) if prio_r else 'NULL'}, 1, "
                f"{sql_escape(detail)}, {sql_escape(flash) if flash else 'NULL'}, {sql_escape(pro) if pro else 'NULL'}, "
                f"{sql_escape(ref_json)}) "
                "ON DUPLICATE KEY UPDATE title=VALUES(title), category_key=VALUES(category_key), shilue_kind=VALUES(shilue_kind), "
                "blurb=VALUES(blurb), detail_md=VALUES(detail_md), detail_md_flash=VALUES(detail_md_flash), detail_md_pro=VALUES(detail_md_pro), "
                "original_ref_json=VALUES(original_ref_json), business_code=VALUES(business_code);"
            )
            excel_box_ids.append(bid)

    # 同一批史略 ID 的子表先清空，避免重复执行脚本时评述/见证/关系图倍增
    if excel_box_ids:
        ids_sql = ",".join(sql_escape(b) for b in sorted(set(excel_box_ids)))
        emit(f"DELETE FROM box_graph_edge WHERE box_id IN ({ids_sql});")
        emit(f"DELETE FROM box_graph_node WHERE box_id IN ({ids_sql});")
        emit(f"DELETE FROM box_critique WHERE box_id IN ({ids_sql});")
        emit(f"DELETE FROM box_relic WHERE box_id IN ({ids_sql});")

    # --- 评述 ---
    crit_sort: dict[str, int] = defaultdict(int)
    if "评述" in wb.sheetnames:
        h, data = sheet_rows(wb["评述"])
        for row in data:
            if not row or not row[0]:
                continue
            ctitle = str(row[0] or "").strip()
            shilue_name = str(row[1] or "").strip()
            shilue_id = str(row[2] or "").strip()
            author = str(row[3] or "").strip()
            src_book = str(row[4] or "").strip()
            content = str(row[5] or "").strip()
            blurb = str(row[6] or "").strip() if len(row) > 6 and row[6] else None
            era = str(row[7] or "").strip() if len(row) > 7 else ""
            bid = shilue_id if shilue_id else None
            if not bid:
                continue
            crit_sort[bid] += 1
            emit(
                "INSERT INTO box_critique (box_id, title, author, era_text, year_value, content, source, blurb, sort_order) "
                f"VALUES ({sql_escape(bid)}, {sql_escape(ctitle) if ctitle else 'NULL'}, {sql_escape(author)}, {sql_escape(era)}, NULL, "
                f"{sql_escape(content)}, {sql_escape(src_book)}, {sql_escape(blurb) if blurb else 'NULL'}, {crit_sort[bid]});"
            )

    # --- 见证 ---
    relic_sort: dict[str, int] = defaultdict(int)
    if "见证" in wb.sheetnames:
        h, data = sheet_rows(wb["见证"])
        for row in data:
            if not row or not row[0]:
                continue
            name = str(row[0] or "").strip()
            shilue_id = str(row[2] or "").strip() if len(row) > 2 else ""
            museum = str(row[3] or "").strip() if len(row) > 3 else ""
            desc = str(row[4] or "").strip() if len(row) > 4 else ""
            summ = str(row[5] or "").strip() if len(row) > 5 else ""
            img = str(row[6] or "").strip() if len(row) > 6 else ""
            pcode = str(row[7] or "").strip() if len(row) > 7 else None
            preason = str(row[8] or "").strip() if len(row) > 8 else None
            if not shilue_id:
                continue
            relic_sort[shilue_id] += 1
            img_sql = "NULL" if not img or "待查" in img else sql_escape(img)
            emit(
                "INSERT INTO box_relic (box_id, name, image_url, summary, description, museum, priority_code, priority_reason, sort_order) "
                f"VALUES ({sql_escape(shilue_id)}, {sql_escape(name)}, {img_sql}, {sql_escape(summ) if summ else 'NULL'}, "
                f"{sql_escape(desc)}, {sql_escape(museum)}, {sql_escape(pcode) if pcode else 'NULL'}, "
                f"{sql_escape(preason) if preason else 'NULL'}, {relic_sort[shilue_id]});"
            )

    # --- 关系 -> graph nodes/edges ---
    if "关系" in wb.sheetnames:
        h, data = sheet_rows(wb["关系"])
        by_box: dict[str, list] = defaultdict(list)
        for row in data:
            if not row or len(row) < 4:
                continue
            rid = str(row[0] or "").strip()
            bid = str(row[2] or "").strip()
            node_title = str(row[3] or "").strip()
            if not bid or not rid:
                continue
            cat = str(row[4] or "").strip() if len(row) > 4 else ""
            level = str(row[5] or "").strip() if len(row) > 5 else ""
            edge_lbl = str(row[6] or "").strip() if len(row) > 6 else ""
            l1 = str(row[7] or "").strip() if len(row) > 7 else ""
            l2 = str(row[8] or "").strip() if len(row) > 8 else ""
            l3 = str(row[9] or "").strip() if len(row) > 9 else ""
            l4 = str(row[10] or "").strip() if len(row) > 10 else ""
            summ = str(row[11] or "").strip() if len(row) > 11 else ""
            extra = {
                "relationId": rid,
                "category": cat,
                "level": level,
                "l1": l1,
                "l2": l2,
                "l3": l3,
                "l4": l4,
                "summary": summ,
            }
            nk = re.sub(r"[^a-zA-Z0-9_]+", "_", rid)[:60]
            by_box[bid].append((nk, node_title, cat, edge_lbl, extra))

        for bid, rows in by_box.items():
            root_key = "rel_root"
            emit(
                f"INSERT INTO box_graph_node (box_id, node_key, node_type, name, extra_json) VALUES ({sql_escape(bid)}, {sql_escape(root_key)}, 'root', '根', '{{}}') "
                "ON DUPLICATE KEY UPDATE name=VALUES(name);"
            )
            parent_map: dict[str, str] = {}
            for nk, title, cat, edge_lbl, extra in rows:
                et = "person" if "家庭" in cat or "君臣" in cat else "event"
                ej = json.dumps(extra, ensure_ascii=False)
                emit(
                    "INSERT INTO box_graph_node (box_id, node_key, node_type, name, extra_json) "
                    f"VALUES ({sql_escape(bid)}, {sql_escape(nk)}, {sql_escape(et)}, {sql_escape(title)}, {sql_escape(ej)}) "
                    "ON DUPLICATE KEY UPDATE name=VALUES(name), extra_json=VALUES(extra_json);"
                )
                chain = [x for x in [extra.get("l1"), extra.get("l2"), extra.get("l3"), extra.get("l4")] if x]
                parent_title = chain[-2] if len(chain) >= 2 else (chain[0] if len(chain) == 1 else None)
                pk = root_key
                if parent_title:
                    for pn, pt, _, _, ex in rows:
                        if pt == parent_title:
                            pk = pn
                            break
                lbl = edge_lbl or "关系"
                emit(
                    "INSERT INTO box_graph_edge (box_id, from_node_key, to_node_key, label) "
                    f"VALUES ({sql_escape(bid)}, {sql_escape(pk)}, {sql_escape(nk)}, {sql_escape(lbl[:32])});"
                )

    sql_text = "\n".join(stmts)
    if args.sql_out:
        args.sql_out.write_text(sql_text, encoding="utf-8")
        print("Wrote", args.sql_out, "statements:", len(stmts))

    if args.mysql_host and args.mysql_user and args.mysql_db:
        import pymysql

        conn = pymysql.connect(
            host=args.mysql_host,
            port=args.mysql_port,
            user=args.mysql_user,
            password=args.mysql_password,
            database=args.mysql_db,
            charset="utf8mb4",
        )
        try:
            with conn.cursor() as cur:
                for s in stmts:
                    cur.execute(s)
            conn.commit()
        finally:
            conn.close()
        print("Executed", len(stmts), "statements on", args.mysql_db)


if __name__ == "__main__":
    main()
