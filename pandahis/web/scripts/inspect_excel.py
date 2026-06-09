from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


def main() -> None:
    path = Path(r"C:\Users\user\Desktop\历史图谱\pandahis\prd\历史图谱 - 首页数据.xlsx")
    wb = load_workbook(path, read_only=False, data_only=True)
    out = {"sheets": []}
    for name in wb.sheetnames:
        ws = wb[name]
        # ensure dimensions are computed
        _ = ws.calculate_dimension()
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=min(8, ws.max_row), values_only=True):
            rows.append([cell for cell in row[:80]])
        out["sheets"].append(
            {
                "name": name,
                "maxRow": ws.max_row,
                "maxColumn": ws.max_column,
                "previewRows": rows,
            }
        )

    out_path = Path(__file__).with_name("excel_preview.json")
    out_path.write_text(__import__("json").dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {out_path}")


if __name__ == "__main__":
    main()

