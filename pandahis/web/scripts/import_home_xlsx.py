from __future__ import annotations

import json
import math
import re
import hashlib
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def _slug(s: str) -> str:
    s = s.strip()
    s = re.sub(r"\s+", "-", s)
    s = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff\-_]+", "", s)
    return s[:48] if s else "unknown"


# PRD 3.3.2.2 一级地域：固定 id 与色板变量（与 web/src/data/models.ts 一致）
PRD_REGIONS: list[dict[str, str]] = [
    {"id": "huaxia", "name": "华夏", "colorVar": "--c-huaxia"},
    {"id": "chaoxian", "name": "朝鲜", "colorVar": "--c-chaoxian"},
    {"id": "japan", "name": "日本", "colorVar": "--c-japan"},
    {"id": "sea", "name": "东南亚", "colorVar": "--c-sea"},
    {"id": "india", "name": "印度", "colorVar": "--c-india"},
    {"id": "persia", "name": "波斯两河", "colorVar": "--c-persia"},
    {"id": "arab", "name": "阿拉伯", "colorVar": "--c-arab"},
    {"id": "egypt", "name": "埃及", "colorVar": "--c-egypt"},
    {"id": "eeu", "name": "东欧", "colorVar": "--c-eeu"},
    {"id": "medi", "name": "地中海古典", "colorVar": "--c-medi"},
    {"id": "weu", "name": "西欧", "colorVar": "--c-weu"},
    {"id": "wafrica", "name": "西非", "colorVar": "--c-wafrica"},
    {"id": "camer", "name": "中美洲", "colorVar": "--c-camer"},
    {"id": "andes", "name": "安第斯", "colorVar": "--c-andes"},
]

# (region_id, keywords) 优先匹配靠前
_REGION_MATCH: list[tuple[str, tuple[str, ...]]] = [
    ("huaxia", ("华夏", "中华", "中国", "汉地", "中原")),
    ("chaoxian", ("朝鲜", "韩国", "高句丽", "新罗", "百济")),
    ("japan", ("日本", "大和", "江户", "幕府")),
    ("sea", ("东南亚", "南洋", "越南", "泰国", "高棉", "马来", "印尼")),
    ("india", ("印度", "天竺", "莫卧儿")),
    ("persia", ("波斯", "两河", "美索不达米亚", "伊朗")),
    ("arab", ("阿拉伯", "伊斯兰", "哈里发", "奥斯曼")),
    ("egypt", ("埃及", "法老")),
    ("eeu", ("东欧", "罗斯", "斯拉夫", "波兰", "基辅")),
    ("medi", ("地中海", "罗马", "希腊", "拜占庭", "雅典")),
    ("weu", ("西欧", "英法", "英国", "法国", "德意志", "西班牙")),
    ("wafrica", ("西非", "马里", "加纳", "桑海")),
    ("camer", ("中美洲", "玛雅", "阿兹特克", "墨西哥")),
    ("andes", ("安第斯", "印加", "南美")),
]


def map_region_id(civ2: str, civ1: str) -> str:
    text = f"{civ2} {civ1}"
    for rid, keys in _REGION_MATCH:
        if any(k in text for k in keys):
            return rid
    # 数据里常见「亚洲文明」等：默认归入华夏（以中华主体数据为主）
    if "亚洲" in text or "东亚" in text:
        return "huaxia"
    return "huaxia"


def year_label(year: int) -> str:
    return f"前{abs(year)}" if year < 0 else str(year)

def _stable_int(s: str) -> int:
    h = hashlib.sha256(s.encode("utf-8")).digest()
    return int.from_bytes(h[:8], "big", signed=False)

def _pick(seq: list[Any], seed: str) -> Any:
    if not seq:
        raise ValueError("empty seq")
    i = _stable_int(seed) % len(seq)
    return seq[i]

def _clamp(n: int, lo: int, hi: int) -> int:
    return max(lo, min(hi, n))

def _split_tags(s: str) -> list[str]:
    s = (s or "").strip()
    if not s:
        return []
    # supports: "a,b" "a，b" "a; b" "a / b" "a  b"
    parts = re.split(r"[，,;/\s]+", s)
    parts = [p.strip() for p in parts if p and p.strip()]
    # de-dup while keeping order
    out: list[str] = []
    seen = set()
    for p in parts:
        if p not in seen:
            out.append(p)
            seen.add(p)
    return out[:8]

CAT_KEYS = ["junji", "shichen", "minlu", "dianzhi", "shilue"]
CAT_NAMES = {
    "junji": "君纪",
    "shichen": "士臣",
    "minlu": "民录",
    "dianzhi": "典制",
    "shilue": "事略",
}

def _guess_category(box_id: str, tags: list[str], remark: str) -> str:
    text = f"{box_id} {' '.join(tags)} {remark}"
    # heuristic keywords
    if re.search(r"即位|登基|册立|继位|崩|驾崩|禅让|称帝|改元|年号", text):
        return "junji"
    if re.search(r"变法|法|制|律|令|田|赋|税|官制|科举|改革|条令|册|诏", text):
        return "dianzhi"
    if re.search(r"将|相|臣|名臣|宰相|御史|大夫|学者|使者|外交|谋|谏", text):
        return "shichen"
    if re.search(r"民|灾|饥|徭|役|起义|暴动|迁徙|风俗|贸易|市", text):
        return "minlu"
    return _pick(CAT_KEYS, box_id)

def _mk_paragraphs(unit: str, emperor: str, dynasty: str, era: str, remark: str, tags: list[str], seed: str) -> list[str]:
    # a "明朝那些事儿" tone but safe, non-assertive where uncertain
    hook = _pick(
        [
            f"{unit}，表面是一个名字，背后却是一段时代的起笔。",
            f"如果把历史当成一张地图，{unit} 就是其中一个坐标点。",
            f"很多故事的开头都很安静，{unit} 也不例外。",
            f"{unit} 往往被一句话概括，但真正耐看的是它的纹理。",
        ],
        seed + ":hook",
    )

    context_bits = []
    if dynasty:
        context_bits.append(dynasty)
    if emperor and emperor != "-":
        context_bits.append(emperor)
    if era and era != "-":
        context_bits.append(f"年号「{era}」")
    context = " · ".join(context_bits) if context_bits else "（时代信息待补充）"

    p2 = f"在 {context} 的语境里，它不像一个孤立事件，更像一枚把叙事拧紧的螺丝：你从这里往前看，会看到铺垫；往后看，会看到连锁反应。"

    p3_core = remark.strip() if remark else "这条内容的细节仍在整理中，但它在整体脉络中占据了一个清晰的位置。"
    if tags:
        tag_line = "、".join(tags[:4])
        p3 = f"{p3_core}（标签：{tag_line}）"
    else:
        p3 = p3_core

    p4 = _pick(
        [
            "把它放进“年份 × 五类分类”的矩阵里，你会发现它并不只是一个点，而是一个入口。",
            "当你沿着关系图谱去看，就能理解为什么某些名字总会反复出现。",
            "跨时空评述不是为了给答案，而是为了给你更多视角去判断。",
            "文物与原文只是“证据的一部分”，更重要的是把证据放回结构里理解。",
        ],
        seed + ":p4",
    )

    return [hook, p2, p3, p4]

def _mk_critiques(unit: str, seed: str, importance: int) -> list[dict[str, str]]:
    # generate 1-3 critiques based on importance
    n = 1 if importance <= 3 else (2 if importance == 4 else 3)
    pool = [
        ("史家", "后世 · 评述", f"如果只记住 {unit} 的表象，就会错过它真正推动的结构变化。", "《编者按》"),
        ("学者", "近代 · 评述", f"{unit} 的意义不在于单点成败，而在于它如何重排了利益与秩序。", "《研究札记》"),
        ("旁观者", "当代 · 读法", f"把 {unit} 放回当时的约束条件里理解，结论往往会更克制、更接近真实。", "《读史方法》"),
        ("评论者", "当代 · 讨论", f"{unit} 之所以耐看，是因为它同时触发了权力、制度与人心三条线。", "《讨论摘录》"),
    ]
    out = []
    for i in range(n):
        a, era, content, src = pool[( _stable_int(seed + f":c{i}") ) % len(pool)]
        out.append({"author": a, "eraText": era, "content": content, "source": src})
    return out

def _mk_relics(unit: str, seed: str, importance: int) -> list[dict[str, str]]:
    # PRD：见证最多 3 件
    if importance <= 2:
        n = 0
    elif importance <= 3:
        n = 1
    elif importance == 4:
        n = 2
    else:
        n = 3
    pool = [
        ("相关器物（占位）", f"用于呈现 {unit} 的物质侧面（待补充具体名称、图片与馆藏）。", "博物馆/馆藏待补充"),
        ("文书/诏令（占位）", f"用于呈现 {unit} 的制度侧面（待补充具体条目与出处）。", "馆藏待补充"),
        ("碑刻/铭文（占位）", f"用于呈现 {unit} 的文本证据（待补充具体铭文与释读）。", "馆藏待补充"),
    ]
    out = []
    for i in range(n):
        name, desc, museum = pool[( _stable_int(seed + f":r{i}") ) % len(pool)]
        out.append({"name": name, "description": desc, "museum": museum})
    return out

def _mk_graph(unit: str, emperor: str, dynasty: str, tags: list[str], seed: str) -> dict[str, Any]:
    center = "event_center"
    nodes = [{"key": center, "name": unit or "事件", "kind": "event"}]
    edges = []
    if emperor and emperor != "-" and emperor != unit:
        nodes.append({"key": "person_emperor", "name": emperor, "kind": "person"})
        edges.append({"from": center, "to": "person_emperor", "label": "相关"})
    if dynasty:
        nodes.append({"key": "org_dynasty", "name": dynasty, "kind": "org"})
        edges.append({"from": center, "to": "org_dynasty", "label": "时代"})
    for i, t in enumerate(tags[:3]):
        k = f"tag_{i}"
        nodes.append({"key": k, "name": t, "kind": "work" if i == 0 else "place"})
        edges.append({"from": center, "to": k, "label": "关联"})
    return {"center": center, "nodes": nodes, "edges": edges}

def _topic_id(title: str) -> str:
    return _slug(title)

def _mk_topics(units: list[dict[str, Any]], boxes: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """
    Build shareable topics from existing content.
    Rules (v1):
    - Topic candidates from: dynasty, unit name, and top tags.
    - Each topic collects related boxes by keyword match in title/detail.
    - Keep only topics with enough matches, and pick hero items.
    """
    text_by_box: dict[str, str] = {}
    for b in boxes:
        text_by_box[b["id"]] = " ".join([b.get("title", ""), " ".join(b.get("detail", []))])

    # collect candidate keywords
    cand: list[str] = []
    for u in units:
        name = (u.get("name") or "").strip()
        if name and name not in cand:
            cand.append(name)
        for part in str(u.get("crumbText") or "").split(" · "):
            part = part.strip()
            if part and part not in cand and len(part) <= 12:
                cand.append(part)
    # tags from boxes' graph nodes (we used tags as nodes sometimes) + details
    # (keep it light: only first N unique short tokens that look like Chinese words)
    for b in boxes:
        for token in re.findall(r"[\u4e00-\u9fff]{2,6}", " ".join(b.get("detail", []))):
            if token not in cand:
                cand.append(token)

    cand = cand[:180]

    topics: list[dict[str, Any]] = []
    for kw in cand:
        matches = [bid for bid, t in text_by_box.items() if kw in t]
        if len(matches) < 3:
            continue
        # pick hero boxes (by importance)
        box_map = {b["id"]: b for b in boxes}
        hero_boxes = sorted(matches, key=lambda bid: (-int(box_map[bid].get("importanceLevel", 3)), bid))[:8]
        hero_units = []
        for bid in hero_boxes:
            uid = box_map[bid].get("unitId")
            if uid and uid not in hero_units:
                hero_units.append(uid)
        title = kw
        subtitle = f"从「{kw}」出发，沿着盒子与关系图谱向外延伸。"
        topics.append(
            {
                "id": _topic_id(title),
                "title": title,
                "subtitle": subtitle,
                "keywords": [kw],
                "heroBoxIds": hero_boxes,
                "heroUnitIds": hero_units[:6],
            }
        )

    # de-dup by id and cap
    uniq: dict[str, dict[str, Any]] = {}
    for t in topics:
        if t["id"] not in uniq:
            uniq[t["id"]] = t
    out = list(uniq.values())
    out.sort(key=lambda t: (-len(t.get("heroBoxIds", [])), t["title"]))
    return out[:60]


@dataclass(frozen=True)
class Row:
    box_id: str
    dynasty: str
    emperor: str
    unit_name: str
    era: str
    start_year: int
    end_year: int
    duration_years: int
    importance: int
    remark: str
    civ2: str
    tags: str
    status: str
    civ1: str


def parse_units_and_boxes(xlsx_path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """从 PRD Excel 解析 units / boxes（与 main 写入 generated.json 的结构一致，并含入库所需字段）。"""
    wb = load_workbook(xlsx_path, read_only=False, data_only=True)

    def find_dynasty_sheet() -> Any:
        for name in wb.sheetnames:
            sh = wb[name]
            hdr = [str(c).strip() if c is not None else "" for c in next(sh.iter_rows(min_row=1, max_row=1, values_only=True))]
            if "历史盒子 ID" in hdr:
                return sh, hdr
        raise KeyError("未找到含「历史盒子 ID」的工作表（新版 PRD 一般为「朝代」表）")

    ws, headers = find_dynasty_sheet()
    idx = {h: i for i, h in enumerate(headers)}

    def g(row: list[Any], key: str, default: Any = "") -> Any:
        i = idx.get(key)
        if i is None or i >= len(row):
            return default
        v = row[i]
        return default if v is None else v

    rows: list[Row] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        row = list(raw)
        box_id = str(g(row, "历史盒子 ID", "")).strip()
        if not box_id:
            continue
        try:
            start_year = int(g(row, "即位时间", 0))
            end_year = int(g(row, "退位时间", start_year))
        except Exception:
            continue

        rows.append(
            Row(
                box_id=box_id,
                dynasty=str(g(row, "朝代名称", "")).strip(),
                emperor=str(g(row, "帝王名称", "")).strip(),
                unit_name=str(g(row, "历史单元名称", "")).strip(),
                era=str(g(row, "年号", "")).strip(),
                start_year=start_year,
                end_year=end_year,
                duration_years=int(g(row, "在位时长", max(0, end_year - start_year))),
                importance=int(g(row, "重要性评级", 3)),
                remark=str(g(row, "备注", "")).strip(),
                civ2=str(g(row, "文明体系", "")).strip(),
                tags=str(g(row, "标签", "")).strip(),
                status=str(g(row, "历史盒子状态", "")).strip(),
                civ1=str(g(row, "一级文明体系", "")).strip(),
            )
        )

    civilizations = [{"id": r["id"], "name": r["name"], "colorVar": r["colorVar"]} for r in PRD_REGIONS]

    unit_key_to_unit_id: dict[tuple[str, str, str, str, str], str] = {}
    units_by_id: dict[str, dict[str, Any]] = {}
    boxes: list[dict[str, Any]] = []

    for r in rows:
        civ_id = map_region_id(r.civ2, r.civ1)
        civ_name = next((c["name"] for c in civilizations if c["id"] == civ_id), PRD_REGIONS[0]["name"])

        k = (civ_id, r.unit_name, r.dynasty, r.emperor, r.era)
        unit_id = unit_key_to_unit_id.get(k)
        if not unit_id:
            unit_id = _slug(f"{civ_id}-{r.unit_name}-{r.dynasty}-{r.emperor}-{r.era}")
            unit_key_to_unit_id[k] = unit_id

            crumb_parts = [civ_name]
            if r.dynasty and r.dynasty != "-":
                crumb_parts.append(r.dynasty)
            if r.emperor and r.emperor != "-" and r.emperor != r.unit_name:
                crumb_parts.append(r.emperor)
            crumb_text = " · ".join([p for p in crumb_parts if p])

            dyn = r.dynasty if r.dynasty and r.dynasty != "-" else ""

            units_by_id[unit_id] = {
                "id": unit_id,
                "name": r.unit_name or r.emperor or r.dynasty or r.box_id,
                "dynastyName": dyn,
                "civilizationId": civ_id,
                "crumbText": crumb_text,
                "eraText": f"年号 {r.era}" if r.era and r.era != "-" else None,
                "startYear": r.start_year,
                "endYear": r.end_year,
                "durationYears": r.duration_years,
                "summary": r.remark or "（待补充）",
            }
        else:
            u = units_by_id[unit_id]
            u["startYear"] = min(u["startYear"], r.start_year)
            u["endYear"] = max(u["endYear"], r.end_year)
            u["durationYears"] = max(u["durationYears"], r.duration_years)
            if r.remark and (not u.get("summary") or u["summary"] == "（待补充）"):
                u["summary"] = r.remark
            dyn = r.dynasty if r.dynasty and r.dynasty != "-" else ""
            if dyn and not u.get("dynastyName"):
                u["dynastyName"] = dyn

        tags = _split_tags(r.tags)
        category_key = _guess_category(r.box_id, tags, r.remark)
        title = r.unit_name or r.emperor or r.box_id
        seed = r.box_id

        boxes.append(
            {
                "id": r.box_id,
                "unitId": unit_id,
                "title": title,
                "categoryKey": category_key,
                "year": r.start_year,
                "startYear": r.start_year,
                "endYear": r.end_year,
                "importanceLevel": _clamp(r.importance, 1, 5),
                "excelStatus": r.status,
                "detail": _mk_paragraphs(r.unit_name, r.emperor, r.dynasty, r.era, r.remark, tags, seed),
                "critiques": _mk_critiques(title, seed, _clamp(r.importance, 1, 5)),
                "relics": _mk_relics(title, seed, _clamp(r.importance, 1, 5)),
                "graph": _mk_graph(title, r.emperor, r.dynasty, tags, seed),
            }
        )

    return list(units_by_id.values()), boxes


def main() -> None:
    repo_root = Path(__file__).resolve().parent.parent.parent
    candidates = [
        repo_root / "prd" / "历史图谱.xlsx",
        repo_root / "prd" / "历史图谱 - 首页数据.xlsx",
    ]
    xlsx_path = next((p for p in candidates if p.is_file()), None)
    if not xlsx_path:
        raise FileNotFoundError(f"未找到 Excel，尝试过: {[str(p) for p in candidates]}")

    units, boxes = parse_units_and_boxes(xlsx_path)

    # time axis: use units start years, but keep it within a comfortable number for rendering
    years = sorted({int(u["startYear"]) for u in units})
    target = 28
    if len(years) > target:
        step = max(1, math.floor(len(years) / target))
        years = years[::step]
    time_axis = [{"year": y, "label": year_label(int(y)), "rowHeightPx": 56} for y in years]

    search_suggest = {
        "hot": [{"keyword": k, "hot": True} for k in ["伏羲", "女娲", "神农", "黄帝", "夏禹治水", "商汤革命"]],
        "history": ["三皇五帝", "文化繁荣", "统一全国"],
    }

    topics = _mk_topics(units, boxes)

    civilizations = [{"id": r["id"], "name": r["name"], "colorVar": r["colorVar"]} for r in PRD_REGIONS]

    out = {
        "civilizations": civilizations,
        "timeAxis": time_axis,
        "units": units,
        "boxes": boxes,
        "searchSuggest": search_suggest,
        "topics": topics,
    }

    out_path = repo_root / "web" / "src" / "data" / "generated.json"
    out_path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {out_path} (from {xlsx_path})")


if __name__ == "__main__":
    main()

